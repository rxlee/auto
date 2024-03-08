import json
import os
import time
from datetime import datetime
from openpyxl import load_workbook
import random


time.sleep(random.randint(5, 10))

now = int(time.time())
today = time.strftime("%Y%m%d", time.localtime())
cappath = 'cap/' + str(today)


wb = load_workbook('每日数据.xlsx')
ws_bsr = wb['排名']

last = ws_bsr.cell(row=1, column=ws_bsr.max_column).value
write_col = ws_bsr.max_column
if today != last:  # 新的一天，往后新写一列
    write_col += 1
ws_bsr.cell(1, write_col).value = today

# 读取船长销量数据
data = {}
if os.path.exists(cappath):
    data = json.load(open(cappath))
else:
    exit()

# 初始化ASIN销量关系
asin_unit_sell = {}
keywords = {}
for sku in data:
    if sku == 'time':
        continue
    if sku == 'keywords':
        keywords = data[sku]
        continue
    # print(sku+'\t'+str(data[sku]['_1_day_sale'])+'\t'+str(data[sku]['yesterday_day_sale']))
    asin_unit_sell[data[sku]['asin']] = data[sku]['_1_day_sale']



for i in range(2, ws_bsr.max_row + 1):
    asin = ws_bsr.cell(i, 2).value
    if asin is None or asin.strip() == '':
        continue
    type = ws_bsr.cell(i, 3).value
    json_file = 'json/'+asin+'.json'
    js = {}
    if type in ('BSR', 'UNIT'):
        if os.path.exists(json_file):
            js = json.load(open(json_file))
        else:
            print('asin数据消失：'+asin)
            continue
    if type == 'BSR':
        st_today_start = int(now/60/60/24)
        for j in js['change_bsr']:
            ws_bsr.cell(i, write_col).value = '未更新'
            if int(j['time']/60/60/24) >= st_today_start:
                ws_bsr.cell(i, write_col).value = j['bsr']
                break
    if type == 'UNIT':
        ws_bsr.cell(i, write_col).value = asin_unit_sell[asin]
    if type == 'KEYWORD' and asin in keywords:  # asin means keyword here
        ws_bsr.cell(i, write_col).value = keywords[asin]

wb.save('每日数据.xlsx')