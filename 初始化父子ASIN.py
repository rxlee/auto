from openpyxl import load_workbook
import os
import time
import json

asin_to_pasin = {}

# jsonpath = 'json/children-to-parent.json'
# # jsonpath = 'json/parent-to-children.json'
#
# if os.path.exists(jsonpath):
#     asin_to_pasin = json.load(open(jsonpath))
#
# print(asin_to_pasin['B07FD31G8M'])


wb = load_workbook('父子ASIN源.xlsx', data_only=True)
ws = wb['sheet']
ptc = {}
ctp = {}
asin_infos = {}
for i in range(1, ws.max_row + 1):
    pasin = ws.cell(i, 1).value
    asin = ws.cell(i, 2).value
    ctp[asin] = pasin
    if pasin not in ptc:
        ptc[pasin] = []
    ptc[pasin].append(asin)
    asin_info = {'sku': ws.cell(i, 3).value, 'name': ws.cell(i, 4).value}
    asin_infos[asin] = asin_info



with open('json/parent-to-children.json', 'w') as fp:
    json.dump(ptc, fp)

with open('json/children-to-parent.json', 'w') as fp:
    json.dump(ctp, fp)

with open('json/asin-info.json', 'w') as fp:
    json.dump(asin_infos, fp)
