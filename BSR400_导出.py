from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import time
import os
from os import path
import random

dir_json = 'bsr400'

wb_target = load_workbook('bsr400.xlsx')

def to_excel(f):
    ws = wb_target.create_sheet(f.replace(dir_json, '').replace('.json', '').replace('\\', ''))
    data = json.load(open(f))
    ws.cell(1, 1).value = 'asin'
    ws.cell(1, 2).value = 'bsr'
    ws.cell(1, 3).value = '评分'
    ws.cell(1, 4).value = 'rating'
    ws.cell(1, 5).value = '价格'
    ws.cell(1, 6).value = '标题'
    for i in data['data']:
        bsr = int(i['metadataMap']['render.zg.rank'])
        row = bsr + 1
        ws.cell(row, 1).value = i['id']
        ws.cell(row, 2).value = bsr
        if 'bsr' not in i:
            continue
        ws.cell(row, 3).value = i['score']
        ws.cell(row, 4).value = i['rating']
        ws.cell(row, 5).value = i['price']
        ws.cell(row, 6).value = i['title']
        # print(i['id'], i['bsr'], i['metadataMap']['render.zg.rank'])


files = os.listdir(dir_json)
sorted(files)

for f in files:
    # if 'json' not in files:
    #     continue
    real_url = path.join(dir_json, f)
    print(real_url)
    to_excel(real_url)


wb_target.save('C:/Users/Administrator/Desktop/bsr400.xlsx')
