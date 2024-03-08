from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import os
import time
from datetime import datetime
from openpyxl import load_workbook

''' 日常运营数据 '''

week_dic = {
    1: '星期一',
    2: '星期二',
    3: '星期三',
    4: '星期四',
    5: '星期五',
    6: '星期六',
    7: '星期日',
}

now = int(time.time())
today = time.strftime("%Y%m%d", time.localtime())
week_day = week_dic[datetime.now().isoweekday()]
ctp = 'json/children-to-parent.json'

chrome_opt = webdriver.ChromeOptions()
chrome_opt.add_experimental_option("excludeSwitches", ["ignore-certificate-errors"])
chrome_opt.add_argument('--disable-gpu')
chrome_opt.add_argument('--allowed-origins')
service = webdriver.chrome.service.Service(r"C:/Users/Administrator/chromedriver.exe")
br = webdriver.Chrome(service=service, options=chrome_opt)
br.get('https://www.amazon.com/s?me=A1KU19I5STQO9H&marketplaceID=ATVPDKIKX0DER')
time.sleep(1)

asin_to_pasin = {}

if os.path.exists(ctp):
    asin_to_pasin = json.load(open(ctp))

res_rating = {}
nodes = br.find_elements(By.XPATH, '//*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div')
for idx, i in enumerate(nodes):
    try:
        asin = br.find_element(By.XPATH, '//*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[' + str(
            idx) + ']').get_attribute(
            'data-asin')
        if asin == '':
            continue
        score = br.find_element(By.XPATH, '//*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[' + str(idx) +
                                ']/div/div/div/div/div/div[2]/div/div/div[2]/div/span[1]').get_attribute(
            'aria-label').split()[0]
        rating = br.find_element(By.XPATH,
                                 '//*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[' + str(
                                     idx) +
                                 ']/div/div/div/div/div/div[2]/div/div/div[2]/div/span[2]').get_attribute(
            'aria-label').replace(',', '')
        pasin = asin_to_pasin[asin]
        res_rating[pasin] = str(score) + '+' + str(rating)
    except:
        continue

print(res_rating)
wb = load_workbook('每日数据.xlsx')
ws_rating = wb['评分']

last = ws_rating.cell(row=1, column=ws_rating.max_column).value
write_col = ws_rating.max_column
if today != last:  # 新的一天，往后新写一列
    write_col += 1
ws_rating.cell(1, write_col).value = today
ws_rating.cell(2, write_col).value = week_day

for i in range(3, ws_rating.max_row + 1):
    pasin = ws_rating.cell(i, 2).value
    if pasin in res_rating:
        ws_rating.cell(i, write_col).value = res_rating[pasin]
    else:
        print('父体消失：'+ws_rating.cell(i, 1))

wb.save('每日数据.xlsx')

br.close()
