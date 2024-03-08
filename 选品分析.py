from openpyxl import load_workbook
wb = load_workbook('bak.xlsx')
import time

now = int(time.time())
s3m = 91*24*60*60
s6m = 182*24*60*60
s12m = 365*24*60*60


count = 54

cols = {
    '品牌': 'C',
    '标题': 'D',
    '月销量': 'K',
    '留评率':  'R',
    '评分': 'S',
    '上架时间':  'V',
    '配送方式':  'W',
    '卖家数': 'Y',
    '卖家所属地': 'Z',
    'BuyBox类型': 'AB',
    '价格': 'N',
}

ws = wb.worksheets[0]
datas = []
for i in range(2, count + 2):
    istr = str(i)
    # print(ws['V'+istr].value)
    datas.append({'A': ws['A'+istr].value,
                  'B': ws['B'+istr].value,
                  'C': ws['C'+istr].value,
                  'D': ws['D'+istr].value,
                  'E': ws['E'+istr].value,
                  'F': ws['F'+istr].value,
                  'G': ws['G'+istr].value,
                  'H': ws['H'+istr].value,
                  'I': ws['I'+istr].value,
                  'J': ws['J'+istr].value,
                  'K': ws['K'+istr].value,
                  'L': ws['L'+istr].value,
                  'M': ws['M'+istr].value,
                  'N': ws['N'+istr].value,
                  'O': ws['O'+istr].value,
                  'P': ws['P'+istr].value,
                  'Q': ws['Q'+istr].value,
                  'R': ws['R'+istr].value,
                  'S': ws['S'+istr].value,
                  'T': ws['T'+istr].value,
                  'U': ws['U'+istr].value,
                  'V': ws['V'+istr].value,
                  'W': ws['W'+istr].value,
                  'X': ws['X'+istr].value,
                  'Y': ws['Y'+istr].value,
                  'Z': ws['Z'+istr].value,
                  'AA': ws['AA'+istr].value,
                  'AB': ws['AB'+istr].value,
                  'AC': ws['AC'+istr].value,
                  'AD': ws['AD'+istr].value,
                  'AE': ws['AE'+istr].value,
                  'AF': ws['AF'+istr].value,
                  'AG': ws['AG'+istr].value,
                  'AH': ws['AH'+istr].value,
                  'AI': ws['AI'+istr].value,
                  'AJ': ws['AJ'+istr].value,
                  'create_time': int(time.mktime(time.strptime(ws['V'+istr].value, "%Y-%m-%d"))),

                  }
                 )



def c4gb():
    count_all = 0
    brands = []
    for l in datas:
        title = l[cols['标题']].lower()
        brand = l[cols['品牌']]
        if ' 4gb' in title or ' 4 gb' in title:
            brands.append(brand)
            print(l['A'])
            count_all += 1
    print(count_all,len(list(set(brands))))

c4gb()











def dataDateRange():
    count_all = 0
    all_xl = 0
    all_pf = 0
    all_lp = 0
    all_mj = 0
    all_jg = 0
    for l in datas:
        brand = l[cols['品牌']]
        xl = l[cols['月销量']]
        pf = l[cols['评分']]
        lp = l[cols['留评率']]
        mj = l[cols['卖家数']]
        jg = l[cols['价格']]
        # if price == '':
        #     continue
        btw = now - l['create_time']
        if btw >=s12m:
            if xl != '':
                all_xl += xl
            if pf != '':
                all_pf += pf
            if lp != '':
                all_lp += lp
            if mj != '':
                all_mj += mj
            if jg != '':
                all_jg += jg
            count_all += 1
    print(count_all,',', all_xl,',', all_pf/count_all,',', all_mj,',', all_lp/count_all,',', all_jg/count_all)


# dataDateRange()

def brand():
    result_xl = {}
    result_count = {}
    avg_xl = {}  # 平均listing销量
    for data in datas:
        brand = data[cols['品牌']]
        xl = data[cols['月销量']]
        if xl == '':
            xl = 0
        if brand in result_xl:
            result_xl[brand] = result_xl[brand] + xl
            result_count[brand] += 1
        else:
            result_xl[brand] = xl
            result_count[brand] = 1
    for key in result_xl:
        avg_xl[key] = result_xl[key]/result_count[key]
    result = sorted(avg_xl.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    print(result)
    for i in range(0,10):
        print(result[i][0],result_count[result[i][0]],result_xl[result[i][0]],result[i][1],'\n')
    print(sorted(result_xl.items(), key = lambda kv:(kv[1], kv[0]),reverse=True))
    print(sorted(result_count.items(), key = lambda kv:(kv[1], kv[0]),reverse=True))

# 找出靠前的品牌
# brand()



def dataBrand(march):
    count_3m_in = 0
    count_6m_3m = 0
    count_12m_6m = 0
    count_12m_out = 0
    count_all = 0
    all_xl = 0
    all_pf = 0
    all_lp = 0
    all_mj = 0
    all_jg = 0
    for l in datas:
        title = l[cols['标题']].lower()
        brand = l[cols['品牌']]
        xl = l[cols['月销量']]
        pf = l[cols['评分']]
        lp = l[cols['留评率']]
        mj = l[cols['卖家数']]
        jg = l[cols['价格']]
        # if price == '':
        #     continue
        march = str(march)
        if 'class 10' in title:
            if xl != '':
                all_xl += xl
            if pf != '':
                all_pf += pf
            if lp != '':
                all_lp += lp
            if mj != '':
                all_mj += mj
            if jg != '':
                all_jg += jg
            count_all += 1
            btw = now - l['create_time']
            if btw < s3m:
                count_3m_in += 1
            elif btw < s6m:
                count_6m_3m += 1
            elif btw < s12m:
                count_12m_6m += 1
            else:
                count_12m_out += 1
    print(march,',', count_all,',', all_xl,',', all_pf/count_all,',', all_mj,',', all_lp/count_all,',', all_jg/count_all,',', count_3m_in,',', count_6m_3m,',', count_12m_6m,',', count_12m_out)

# dataBrand('uhs-i ')

# dataBrand('SanDisk')
# dataBrand('SAMSUNG')
# dataBrand('Gigastone')
# dataBrand('INLAND')
# dataBrand('alisinsen')
# dataBrand('nuiflash')
# dataBrand('PNY')
# dataBrand('Kingston')





def dateRangeCount(march):
    count_3m_in = 0
    count_6m_3m = 0
    count_12m_6m = 0
    count_12m_out = 0
    count_all = 0
    # price_3m_in = 0
    # price_6m_3m = 0
    # price_12m_6m = 0
    # price_12m_out = 0
    price_all = 0
    for l in datas:
        title = l[cols['标题']]
        price = l[cols['价格']]
        if price == '':
            continue
        if (march != '1T' and march in title) or (march == '1T' and (march in title or '1024GB' in title)) or march == 'ALL':
            count_all += 1
            price_all += price
            btw = now - l['create_time']
            if btw < s3m:
                count_3m_in += 1
                # price_3m_in += price
            elif btw < s6m:
                count_6m_3m += 1
                # price_6m_3m += price
            elif btw < s12m:
                count_12m_6m += 1
                # price_12m_6m += price
            else:
                count_12m_out += 1
                # price_12m_out += price
    print(march, price_all/count_all)

# 容量上线时间
# dateRangeCount('16GB')
# dateRangeCount('32GB')
# dateRangeCount('64GB')
# dateRangeCount('128GB')
# dateRangeCount('256GB')
# dateRangeCount('512GB')
# dateRangeCount('1T')
# dateRangeCount('ALL')



def dateSDXC():
    count_3m_in = 0
    count_6m_3m = 0
    count_12m_6m = 0
    count_12m_out = 0
    count_all = 0
    for l in datas:
        title = l[cols['标题']].lower()
        price = l[cols['价格']]
        if price == '':
            continue
        if 'sdhc' in title:
            count_all += 1
            btw = now - l['create_time']
            if btw < s3m:
                count_3m_in += 1
                # price_3m_in += price
            elif btw < s6m:
                count_6m_3m += 1
                # price_6m_3m += price
            elif btw < s12m:
                count_12m_6m += 1
                # price_12m_6m += price
            else:
                count_12m_out += 1
                # price_12m_out += price
    print('SDXC', count_all,count_3m_in,count_6m_3m,count_12m_6m,count_12m_out)

# dateSDXC()


def dataPackCount(march):
    count_3m_in = 0
    count_6m_3m = 0
    count_12m_6m = 0
    count_12m_out = 0
    count_all = 0
    price_all = 0
    for l in datas:
        title = l[cols['标题']].lower()
        price = l[cols['价格']]
        # if price == '':
        #     continue
        march = str(march)
        # if (march+' pack' in title or march+'-pack' in title) or march == '0':
        if ('pack' in title or 'bulk' in title) and march == '0':
            count_all += 1
            price_all += price
            btw = now - l['create_time']
            if btw < s3m:
                count_3m_in += 1
                # price_3m_in += price
            elif btw < s6m:
                count_6m_3m += 1
                # price_6m_3m += price
            elif btw < s12m:
                count_12m_6m += 1
                # price_12m_6m += price
            else:
                count_12m_out += 1
                # price_12m_out += price
    print(march, count_all, price_all/count_all,count_3m_in,count_6m_3m,count_12m_6m,count_12m_out)

# dataPackCount(2)
# dataPackCount(3)
# dataPackCount(5)
# dataPackCount(10)
# dataPackCount(20)
# dataPackCount(50)
# dataPackCount(100)
# dataPackCount(0)


def dataShippingCount(march):
    count_3m_in = 0
    count_6m_3m = 0
    count_12m_6m = 0
    count_12m_out = 0
    count_all = 0
    for l in datas:
        shipping = l[cols['配送方式']]
        price = l[cols['价格']]
        march = str(march)
        if shipping == march:
            count_all += 1
            btw = now - l['create_time']
            if btw < s3m:
                count_3m_in += 1
                # price_3m_in += price
            elif btw < s6m:
                count_6m_3m += 1
                # price_6m_3m += price
            elif btw < s12m:
                count_12m_6m += 1
                # price_12m_6m += price
            else:
                count_12m_out += 1
                # price_12m_out += price
    # print(march, count_all, price_all/count_all,count_3m_in,count_6m_3m,count_12m_6m,count_12m_out)

# dataShippingCount("AMZ")
# dataShippingCount("FBA")
# dataShippingCount("FBM")
# dataShippingCount("NA")