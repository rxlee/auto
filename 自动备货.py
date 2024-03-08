from openpyxl import load_workbook
import os
import time
import json

'''
备货前准备和注意事项：
1-记得更新《库存记录表.xlsx》中的“申请中”数据；
2-有过新的货件上传动作需要更新FBA_gather，并更新申请中数据
'''
# type运输方式：0-空运；1-海运
type = 0
DAYS_PRE = 50 # 空运备货天数(包含了申请中+在途等)
DAYS_PRE_OC = 6  # 海运备货天数（只跟销量相关，与当前库存无关）

today = time.strftime("%Y-%m-%d", time.localtime())
jsonpath = 'cap/' + str(today).replace('-', '')

ex_ds = load_workbook('库存记录表.xlsx')
ex_tg = load_workbook('备货模板表.xlsx')

# 计算最近30天，断货天数，乘以系数
ratios_oos = {}
st_ds_ratio = ex_ds['容差系数']
for i in range(2, st_ds_ratio.max_row + 1):
    ratio = float(st_ds_ratio.cell(i, 2).value)
    ratios_oos[st_ds_ratio.cell(i, 1).value] = ratio

# 计算30天内的在售天数----减去的断货天数已经乘了容差系数
days_sell = {}
st_ds_oos = ex_ds['断货记录']
oos_max_col = st_ds_oos.max_column
print('断货天数统计周期：从【'+st_ds_oos.cell(1, oos_max_col-29).value+'】至【'+st_ds_oos.cell(1, oos_max_col).value+'】')
for i in range(2, st_ds_oos.max_row + 1):
    day_oos = 0
    for j in range(oos_max_col-29, oos_max_col + 1):
        day_oos += st_ds_oos.cell(i, j).value
    sku = st_ds_oos.cell(i, 1).value
    ratio = 0.8 if sku not in ratios_oos else ratios_oos.get(sku)
    days_sell[sku] = 30 - (30 - day_oos) * ratio
    # print(sku+'\t'+str(day_oos)+'\t'+str(ratio)+'\t'+ str(days_sell[sku]))

# 计算申请中的货件数量
requestings = {}
st_ds_req = ex_ds['申请中']
for i in range(2, st_ds_req.max_row + 1):
    sku = st_ds_req.cell(i, 1).value
    if sku is None or sku.strip() == '':
        continue
    n = st_ds_req.cell(i, 2).value
    if n is None or str(n).strip() == '':
        continue
    num = n if sku not in requestings else requestings[sku] + n
    requestings[sku] = num

# for i in requestings:
#     print(i, requestings[i])

# 获取最新一列的数据
def latest(day):
    latest = {}
    st = ex_ds[str(day)]
    for i in range(2, st.max_row + 1):
        latest[st.cell(i, 1).value] = st.cell(i, st.max_column).value
    return latest


sell_7 = latest(7)
sell_14 = latest(14)
sell_30 = latest(30)
FBA_total = latest('total')
FBA_unsellable = latest('不可售')

# 备货方案一 S30+(S7+S14)/4
res_1 = {}
sell_month = {}
for sku in sell_30:
    sell_month[sku] = sell_30[sku] + (sell_7[sku]+sell_14[sku]) / 4
    if type == 0:  # 空运
        reqesting = 0 if sku not in requestings else requestings[sku]
        res_1[sku] = 0 if days_sell[sku] == 0 else round(DAYS_PRE * sell_month[sku] / days_sell[sku]) - FBA_total[sku] - reqesting + FBA_unsellable[sku]
    elif type == 1:  # 海运备货量
        res_1[sku] = 0 if days_sell[sku] == 0 else round(DAYS_PRE_OC * sell_month[sku] / days_sell[sku])
    # if res_1[sku] > -1:
    #     print(sku, sell_month, FBA_total[sku], reqesting, res_1[sku])




res_1_sorted = sorted(res_1.items(), key=lambda x: x[1], reverse=True)

row_current = 10
st_tg_1 = ex_tg['第一票']
for item in res_1_sorted:
    sku = item[0]
    res = item[1]
    if res < 1:
        break
    st_tg_1.cell(row_current, 1).value = sku
    st_tg_1.cell(row_current, 7).value = res
    st_tg_1.cell(row_current, 9).value = sell_month[sku]  # 月销
    reqesting = 0 if sku not in requestings else requestings[sku]
    st_tg_1.cell(row_current, 10).value = reqesting  # 申请中
    st_tg_1.cell(row_current, 11).value = FBA_total[sku] - FBA_unsellable[sku]  # 库存加在途
    st_tg_1.cell(row_current, 12).value = days_sell[sku]  # 在售天数
    row_current += 1

st_tg_1.cell(4, 2).value = today
for i in reversed(range(row_current, st_tg_1.max_row+1)):
    st_tg_1.delete_rows(i)


ex_tg.save('C:/Users/Administrator/Desktop/威科US FBA申请表 '+today+'（x票）.xlsx')



