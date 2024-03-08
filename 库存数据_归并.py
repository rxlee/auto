from openpyxl import load_workbook
import os
import time
import json
import random


time.sleep(random.randint(5, 10))

''' 数据整理和归档 '''

today = time.strftime("%Y%m%d", time.localtime())
jsonpath = 'cap/' + str(today)

wb_target = load_workbook('库存记录表.xlsx')
target = [
    {'sheet': '可售', 'key': 'fulfillable_quantity'},
    {'sheet': '预留', 'key': 'reserved_quantity'},
    {'sheet': '接收', 'key': 'inbound_receiving_quantity'},
    {'sheet': 'shipped', 'key': 'inbound_shipped_quantity'},
    {'sheet': 'working', 'key': 'inbound_working_quantity'},
    {'sheet': 'total', 'key': 'total_quantity'},
    {'sheet': '不可售', 'key': 'unsellable_quantity'},
    {'sheet': '昨日销', 'key': 'yesterday_day_sale'},
    {'sheet': '3', 'key': '_3_day_sale'},
    {'sheet': '7', 'key': '_7_day_sale'},
    {'sheet': '14', 'key': '_14_day_sale'},
    {'sheet': '30', 'key': '_30_day_sale'},
    {'sheet': '60', 'key': '_60_day_sale'},
    {'sheet': '每日退', 'key': '_1_day_sale_return'},
    {'sheet': '3R', 'key': '_3_day_sale_return'},
    {'sheet': '7R', 'key': '_7_day_sale_return'},
    {'sheet': '14R', 'key': '_14_day_sale_return'},
    {'sheet': '30R', 'key': '_30_day_sale_return'},
    {'sheet': '60R', 'key': '_60_day_sale_return'}
]


def update_oos(d):  # 更新断货记录表
    ws_oos = wb_target['断货记录']
    last = ws_oos.cell(row=1, column=ws_oos.max_column).value
    write_col = ws_oos.max_column
    if today != last:  # 新的一天，往后新写一列
        write_col += 1
    ws_oos.cell(row=1, column=write_col).value = today

    for i in range(2, ws_oos.max_row + 1):
        sku = ws_oos.cell(i, 1).value
        if sku in d:
            ws_oos.cell(row=i, column=write_col).value = 1 if d[sku]['fulfillable_quantity'] > 0 else 0
        else:
            ws_oos.cell(row=i, column=write_col).value = 0


def update_sheet(sheet, d, key):
    ws = wb_target[sheet]
    last = ws.cell(row=1, column=ws.max_column).value
    write_col = ws.max_column
    if today != last:  # 新的一天，往后新写一列
        write_col += 1
    ws.cell(row=1, column=write_col).value = today

    for i in range(2, ws.max_row + 1):
        sku = ws.cell(i, 1).value
        if sku in d:
            ws.cell(row=i, column=write_col).value = d[sku][key]
        else:
            ws.cell(row=i, column=write_col).value = 0


data = {}
if os.path.exists(jsonpath):
    data = json.load(open(jsonpath))
else:
    exit()

if len(data) > 0:
    for s in target:  # 提取所有格式化数据
        update_sheet(s.get('sheet'), data, s.get('key'))
    update_oos(data)  # 计算断货与否

wb_target.save('库存记录表.xlsx')
