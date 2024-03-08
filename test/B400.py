from openpyxl import load_workbook

w =  load_workbook('B400.xlsx')
ws = w['US']

caps = {
    '128mb': '128mb',
    '128 mb': '128mb',
    '256mb': '256mb',
    '256 mb': '256mb',
    '512mb': '512mb',
    '512 mb': '512mb',
    '1gb': '1gb',
    '1gb': '1gb',
    '128mb': '128mb',
    '128mb': '128mb',
}

for i in range(2, 401):
    sku = ws.cell(i, 4).value
    title = ws.cell(i, 11).value
    s = (sku + ' ' + title).lower()
    usb = '3.0'
    if '2.0' in s:
        usb = '2.0'
    capacity = ''
    s = ''.join(s.split())
    print(s)
    for ca in ['128mb','256mb','512mb','512gb','256gb','128gb','64gb','32gb','16gb','8gb','4gb','2gb','1gb']:
        if ca in s:
            capacity = ca
            break
    pack = ''
    for p in ['100pack', '50pack', '20pack', '10pack', '5pack', '3pack', '2pack']:
        if p in s:
            pack = p
            break
    ws.cell(i, 8).value = usb
    ws.cell(i, 9).value = capacity
    ws.cell(i, 10).value = pack


w.save('B400.xlsx')