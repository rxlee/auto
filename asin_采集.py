import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import json
import time
import os
import re
import random

# time.sleep(random.randint(2, 4))

dir_json = 'json/'
wb = load_workbook('待监控竞对.xlsx')

now = int(time.time())
t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

t_day = time.strftime("%Y-%m-%d", time.localtime())
ws = wb['ablaze']
asinDict = {}
asinList = []
for i in range(3, ws.max_row+1):
    asin = ws.cell(i, 1).value
    if asin is None or asin.strip() == '':
        continue
    asinList.append(asin)
    asinDict[asin] = {'old':{}, 'new':{}}

print(asinList)
random.shuffle(asinList)

chrome_opt = webdriver.ChromeOptions()
chrome_opt.add_argument('--disable-gpu')
chrome_opt.add_argument('--allowed-origins')
options = Options()
options.add_experimental_option('debuggerAddress', '127.0.0.1:56743')
service = webdriver.chrome.service.Service(r"C:/Users/Administrator/chromedriver.exe")
br = webdriver.Chrome(service=service, options=options)
br.get(
    'https://www.amazon.com/s?k=apple&language=en_US&currency=USD')
time.sleep(1)
try:
    br.find_element(By.ID, 'nav-global-location-popover-link').click()
    time.sleep(4)
    br.find_element(By.XPATH, '//*[@id="GLUXZipUpdateInput"]').send_keys('20005')
    br.find_element(By.XPATH, '//*[@id="GLUXZipUpdateInput"]').send_keys(Keys.ENTER)
    element = br.find_element(By.XPATH, '//*[@id="GLUXConfirmClose"]')
    br.execute_script("arguments[0].click();", element)
    time.sleep(2)
    br.maximize_window()
except:
    pass


def json_update(js, key, value, params={}):
    # if key in js:
    #     value_old = js[key]
    #     if value_old != value:
    if 'change_' + key not in js:
        js['change_' + key] = {'values': []}
    js['change_' + key]['values'].append({'t': now, 'v': value})
    js['change_' + key]['params'] = params
    js[key] = value

re_coupon = re.compile(r'Apply\s(.*?)\scoupon', re.S)  # 最小匹配





def scra(asin):
    print('采集【'+asin+'】...')
    br.get('https://www.amazon.com/dp/' + asin)
    json_file = dir_json + asin + '.json'
    js = {}
    if os.path.exists(json_file):
        js = json.load(open(json_file))
        for k in js:
            asinDict[asin]['old'][k] = js[k]
    status = 'Fine'
    if not br.current_url.split('/')[4].startswith(asin):
        js['status'] = 'Offline'
        return t+'-'+asin + '-缺货下线\n'
    title = 'NONE'
    try:
        title = br.find_element(By.XPATH, '//*[@id="productTitle"]').text
    except:
        return t+'-'+asin + '-标题不见.\n'
    seller = 'NONE'
    try:
        if 'seller' not in js:
            seller = br.find_element(By.XPATH, '//*[@id="sellerProfileTriggerId"]').text
            js['seller'] = seller
    except:
        print(t+'-'+asin + '-Currently unavailable-no seller\n')
        pass
    price = 0.0
    try:
        # price = float(br.find_element(By.XPATH,'//*[@id="centerCol"]').find_elements(By.CLASS_NAME, 'priceToPay')[0].text.replace('\n', '.').replace('$', '').replace(',', ''))
        price = float(br.find_elements(By.CLASS_NAME, 'priceToPay')[0].text.replace('\n', '.').replace('$', '').replace(',', ''))
    except:
        status = 'Currently unavailable.'
        print(t+'-'+asin + '-Currently unavailable-no price\n')
        pass

    list_price = 0.0
    try:
        list_price = float(br.find_element(By.XPATH, '//*[@id="corePriceDisplay_desktop_feature_div"]/div[2]/span/span[1]/span/span/span[1]').text.replace('\n', '.').replace('$', '').replace(',', ''))
    except:
        print(t+'-'+asin + '-no list price\n')
        pass
    prime_price = 0.0
    try:
        prime_price = float(br.find_element(By.XPATH, '//*[@id="pep-signup-link"]/span[2]').text.replace('\n', '.').replace('$', '').replace(',', ''))
    except:
        pass
    rating = 0
    score = 0.0
    # bsr = 0
    bsrs = {}
    bsr_cates = []
    coupon = ''
    deal = ''
    qa = 0
    try:
        qa = int(br.find_element(By.XPATH, '//*[@id="ask_feature_div"]/span[2]').text.split()[0].replace(',', ''))
    except:
        pass

    try:
        trs = br.find_elements(By.XPATH,
                                 '//*[@id="productDetails_detailBullets_sections1"]/tbody/tr')
        for tr in trs:
            try:
                tt = tr.text
                if tt.startswith('Customer Reviews'):
                    ttl = tt.split()
                    rating = int(ttl[3].replace(',', ''))
                    score = float(ttl[2])
                if tt.startswith('Best Sellers Rank'):
                    ttl = tt.replace('Best Sellers Rank ', '').split('\n')
                    for s in ttl:
                        _bsr = s.split()[0]
                        bsr = int(_bsr.replace('#', '').replace(',', ''))
                        cate = s.replace(_bsr + ' in ', '')
                        if s.endswith(')'):
                            cate = cate.split(' (')[0]
                        bsrs[cate] = bsr
                        bsr_cates.append(cate)
                    break
            except:
                continue
    except:
        pass
    try:
        cou = br.find_element(By.XPATH, '//*[@id="ppd_qualifiedBuybox"]/span/div/span/label').text
        coupon = re.findall(re_coupon, cou)[0]
    except:
        pass
    try:
        deal = br.find_element(By.XPATH, '//*[@id="dealBadge_feature_div"]').text
    except:
        pass
    star5p = br.find_element(By.XPATH, '//*[@id="histogramTable"]/tbody/tr[1]').text.split()[2]
    star4p = br.find_element(By.XPATH, '//*[@id="histogramTable"]/tbody/tr[2]').text.split()[2]
    star3p = br.find_element(By.XPATH, '//*[@id="histogramTable"]/tbody/tr[3]').text.split()[2]
    star2p = br.find_element(By.XPATH, '//*[@id="histogramTable"]/tbody/tr[4]').text.split()[2]
    star1p = br.find_element(By.XPATH, '//*[@id="histogramTable"]/tbody/tr[5]').text.split()[2]
    stars = star5p + '-' + star4p + '-' + star3p + '-' + star2p + '-' + star1p
    js['status'] = status
    js['title'] = title
    jsnew = {}
    json_update(js, 'price', price)
    jsnew['price'] = price
    json_update(js, 'list_price', list_price)
    jsnew['list_price'] = list_price
    json_update(js, 'prime_price', prime_price)
    jsnew['prime_price'] = prime_price
    json_update(js, 'qa', qa)
    jsnew['qa'] = qa
    json_update(js, 'score', score)
    jsnew['score'] = score
    json_update(js, 'rating', rating)
    jsnew['rating'] = rating
    json_update(js, 'stars', stars)
    jsnew['stars'] = stars
    json_update(js, 'coupon', coupon)
    jsnew['coupon'] = coupon
    json_update(js, 'deal', deal)
    jsnew['deal'] = deal
    if asin not in js:
        js['asin'] = asin
    if 'bsr_cates' not in js:
        js['bsr_cates'] = []
    for bsr_cate in bsr_cates:
        if bsr_cate not in js['bsr_cates']:
            js['bsr_cates'].append(bsr_cate)
        json_update(js, bsr_cate, bsrs[bsr_cate], {'bsr_cate': bsr_cate})
        jsnew[bsr_cate] = bsrs[bsr_cate]
    asinDict[asin]['new'] = jsnew
    with open(json_file, 'w') as fp:
        json.dump(js, fp)
    return ''

# scra('B08D9PFJB1')
for asin in asinList:
    err = scra(asin)
    if err != '':
        with open('json/error.log', 'a') as f:
            f.write(err)



changed = False
msg = ''

for i in range(3, ws.max_row+1):
    asin = ws.cell(i, 1).value
    if asin not in asinDict:
        continue
    old = asinDict[asin]['old']
    new = asinDict[asin]['new']
    price_o = '' if 'price' not in old else old['price']
    price_n = new['price']
    bsr_o = '' if 'USB Flash Drives' not in old else old['USB Flash Drives']
    bsr_n = new['USB Flash Drives']
    prime_price_o = '' if 'prime_price' not in old else old['prime_price']
    prime_price_n = new['prime_price']
    coupon_o = '' if 'coupon' not in old else old['coupon']
    coupon_n = new['coupon']
    # 价格记录
    ws.cell(i, 7).value = price_n
    if price_n != price_o:
        ws.cell(i, 6).value = price_o
        ws.cell(i, 8).value = price_n - price_o
        msg += (asin+'-的【价格】调整了，从【'+price_o+'】变成了【'+price_n+'】\n')
        changed = True
    # 排名记录
    ws.cell(i, 9).value = bsr_o
    ws.cell(i, 10).value = bsr_n
    ws.cell(i, 11).value = (bsr_n - bsr_o) / bsr_o

    # 会员专享价格记录
    ws.cell(i, 13).value = prime_price_n
    if prime_price_n != prime_price_o:
        ws.cell(i, 12).value = prime_price_o
        ws.cell(i, 14).value = prime_price_n - prime_price_o
        msg += (asin+'-的【Prime专享价格】调整了，从【'+prime_price_o+'】变成了【'+prime_price_n+'】\n')
        changed = True

    # coupon记录
    try:
        ws.cell(i, 16).value = str(coupon_n)
        if coupon_n != coupon_o:
            ws.cell(i, 15).value = coupon_o
            ws.cell(i, 17).value = '有变化'
            msg += (asin+'-的【优惠券】调整了，从【'+coupon_o+'】变成了【'+coupon_n+'】\n')
            changed = True
    except Exception as e:
        print(e)
        print(asin+'coupon 异常-'+coupon_n)
        print(new)
        print(old)

js_changed = {'changed': changed, 'msg': msg}

if changed:
    with open('json/change/'+ t_day +'.json', 'a') as f:
        json.dump(js_changed, f)


wb.save('json/results/' + t_day+ '.xlsx')
wb.close()
