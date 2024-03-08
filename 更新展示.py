from openpyxl import load_workbook
import os
import time
import json
import random


jsonpath = 'html/json/aa.js'
sku_important = []
'''
'''

today = time.strftime("%Y-%m-%d", time.localtime())

wb = load_workbook('库存健康.xlsx')
s = wb['日均销量']
col_key = {}
xAxis = {
    'type': 'category',
    'boundaryGap': False,
    'data': []
  }
for c in range(1, s.max_column + 1):
    head = s.cell(1, c).value
    col_key[c] = head
    if c > 4:
        xAxis['data'].append(head)
print(xAxis['data'])
legend = {
    'data': [],
    'selected': {}
}



series = []
names = {}
for i in range(2, s.max_row + 1):
    sku = s.cell(i, 1).value
    legend['data'].append(sku)
    if sku not in sku_important:
        legend['selected'][sku] = False
    name = s.cell(i, 4).value
    names[sku] = name
    # print(str(sku)+'---'+str(name))
    data = []
    for j in range(5, s.max_column + 1):
        v = s.cell(i, j).value
        data.append('-' if v is None else v)
    serie = {
      'name': sku,
      'type': 'line',
        # 'label':{
        #     'formatter': name,
        #     'show': True
        # },
      'data': data
    }
    series.append(serie)


option = {}
# option['title'] = {'text': '测试'}
option['tooltip'] = {
        'trigger': 'item',
        'axisPointer': {
            'type': 'shadow',
            'label': {
                'backgroundColor': '#6a7985'
            }
        }
      }
option['grid'] = {
    'left': '3%',
    'right': '4%',
    'bottom': '3%',
    'containLabel': True
}
option['toolbox'] = {
    'feature': {
        'saveAsImage': {}
    }
}
option['legend'] = legend
option['xAxis'] = xAxis
option['series'] = series
option['yAxis'] = {'type': 'value'}
option['names'] = names


js = 'var option=' + str(option) + ';'
js = js.replace(': True', ': true')
js = js.replace(': False', ': false')

with open(jsonpath, 'w', encoding='UTF-8') as fp:
    fp.write(js)

os.startfile('html\\unit_day.html')
