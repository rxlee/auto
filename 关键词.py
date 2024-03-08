import traceback

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from telnetlib import EC

import json
import os
import time
import openpyxl
from openpyxl import load_workbook
import re
import random

# 容量查询顺序，因为512GB和2GB同样包含2GB
capacities_search_order = ['128MB','256MB','512MB','512GB','256GB','128GB','64GB','32GB','16GB','8GB','4GB','2GB','1GB','1TB','2TB','1GB']
capacities_words = {
    '128MB': ['128mb', '128 mb' ,'128m'],
    '256MB': ['256mb', '156 mb' ,'156m'],
    '512MB': ['512mb', '512 mb' ,'512m'],
    '1GB': ['1gb', '1 gb' ,'1g'],
    '2GB': ['2gb', '2 gb' ,'2g'],
    '4GB': ['4gb', '4 gb' ,'4g'],
    '8GB': ['8gb', '8 gb' ,'8g'],
    '16GB': ['16gb', '16 gb' ,'16g'],
    '32GB': ['32gb', '32 gb' ,'32g'],
    '64GB': ['64gb', '64 gb' ,'64g'],
    '128GB': ['128gb', '128 gb' ,'128g'],
    '256GB': ['256gb', '256 gb' ,'256g'],
    '512GB': ['512gb', '512 gb' ,'512g'],
    '1TB': ['1tb', '1 tb' ,'1t'],
    '2TB': ['2tb', '2 tb' ,'2t'],
}

pack_search_order = ['2','3','5','10','20','50','100','200']
pack_words = {
    '2': ['2 pack', '2-pack' ,'2pack' ,'2 pcs' ,'2pcs', '2 piece', 'pack of 2'],
    '3': ['3 pack', '3-pack' ,'3pack' ,'3 pcs' ,'3pcs', '3 piece', 'pack of 3'],
    '5': ['5 pack', '5-pack' ,'5pack' ,'5 pcs' ,'5pcs', '5 piece', 'pack of 5'],
    '10': ['10 pack', '10-pack' ,'10pack' ,'10 pcs' ,'10pcs', '10 piece', 'pack of 10'],
    '20': ['20 pack', '20-pack' ,'20pack' ,'20 pcs' ,'20pcs', '20 piece', 'pack of 20'],
    '50': ['50 pack', '50-pack' ,'50pack' ,'50 pcs' ,'50pcs', '50 piece', 'pack of 50'],
    '100': ['100 pack', '100-pack' ,'100pack' ,'100 pcs' ,'100pcs', '100 piece', 'pack of 100'],
    '200': ['200 pack', '200-pack' ,'200pack' ,'200 pcs' ,'200pcs', '200 piece', 'pack of 200'],
}


wb = load_workbook('关键词.xlsx')

now = int(time.time())
t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
ws = wb['备选关键词']
wsb = wb['品牌']
# asinDict = {}
# 比较关注的品牌
brands_concerned = []
for i in range(1, wsb.max_row+1):
    bra = wsb.cell(i, 1).value
    if bra is None or bra.strip() == '':
        continue
    brands_concerned.append(bra.strip())
keywords = []
for i in range(1, ws.max_row+1):
    keyword = {}
    key = ws.cell(i, 1).value
    max_page = ws.cell(i, 2).value
    if key is None or key.strip() == '':
        continue
    keyword['key'] = key
    keyword['max_page'] = max_page
    keywords.append(keyword)

print(keywords)
random.shuffle(keywords)

# random.shuffle(keywords)

dir_keywords = 'keywords/'

now = int(time.time())
t = time.strftime("%Y-%m-%d %H_%M_%S", time.localtime())
file = dir_keywords + str(t) + '.xlsx'

wb = openpyxl.Workbook()

re_coupon = re.compile(r'Save\s(.*?)\swith\scoupon', re.S)  # 最小匹配
# re_coupon = re.compile(r'Save\s(.*)\swith\scoupon', re.S)  # 贪婪匹配
re_low_inventory = re.compile(r'Only\s(.*?)\sleft\sin\sstock', re.S)  # 最小匹配


focus_asins = []
self_asins = {}
if os.path.exists('json/asin-info.json'):
    self_asins = json.load(open('json/asin-info.json'))

options = Options()
# options.add_experimental_option('debuggerAddress', '127.0.0.1:56743')
service = webdriver.chrome.service.Service(r"C:/Users/Administrator/chromedriver.exe")
br = webdriver.Chrome(service=service, options=options)
#
br.get(
    'https://www.amazon.com/s?k=apple&sprefix=apple&language=en_US&currency=USD&ref=nb_sb_noss_2')
time.sleep(1)
try:
    br.find_element(By.ID, 'nav-global-location-popover-link').click()
    time.sleep(4)
    br.find_element(By.XPATH, '//*[@id="GLUXZipUpdateInput"]').send_keys('20005')
    br.find_element(By.XPATH, '//*[@id="GLUXZipUpdateInput"]').send_keys(Keys.ENTER)
    element = br.find_element(By.XPATH, '//*[@id="GLUXConfirmClose"]')
    br.execute_script("arguments[0].click();", element)
    time.sleep(2)
    br.maximize_window()
except:
    pass

cols_list = ['page', 'index', 'asin', 'tags', 'brand', 'score', 'rating', 'price', 'price_old',
             'coupon', 'deal', 'prime_price', 'low_inventory', 'capacity', 'pack','interface', 'original']

pri = 'Join Prime to buy this item at $'.lower()
cols = {}
for i, k in enumerate(cols_list):
    cols[k] = i + 2

total_asins = []

ws_all = wb.create_sheet('汇总')
ws_all.cell(1, 1).value = '搜索词'
for col in cols:
    ws_all.cell(1, cols[col]).value = col

all_line = 1

def scrap(keyword):
    mp = keyword['max_page']
    ky = keyword['key']
    br.find_element(By.XPATH, '//*[@id="twotabsearchtextbox"]').clear()
    br.find_element(By.XPATH, '//*[@id="twotabsearchtextbox"]').send_keys(ky)
    time.sleep(1)
    br.find_element(By.XPATH, '//*[@id="twotabsearchtextbox"]').send_keys(Keys.ENTER)
    time.sleep(1)
    for page in range(1, mp + 1):
        scr_keyword(ky, page)


def scr_keyword(ky, page):
    max_row_all = ws_all.max_row  # 汇总表
    if page > 1:
        try:
            br.find_element(By.CLASS_NAME, 's-pagination-next').click()
        except Exception as e:
            # traceback.print_exc()
            return
    br.execute_script('window.scrollTo(0,3500)')
    time.sleep(0.5)
    br.execute_script('window.scrollTo(0,4500)')
    time.sleep(0.8)
    br.execute_script('window.scrollTo(0,5500)')
    time.sleep(1.2)
    br.execute_script('window.scrollTo(0,6500)')
    time.sleep(1.8)
    nodes = br.find_elements(By.CLASS_NAME, 's-asin')
    idx = 0
    for idx_div, node in enumerate(nodes):
        idx += 1
        try:
            asin = node.get_attribute('data-asin')
            if asin is None or asin == '':
                continue
        except Exception as e:
            traceback.print_exc()
        brand = ''
        try:  # 解析品牌
            brand = node.find_element(By.CLASS_NAME, 's-line-clamp-1').text
        except Exception as e:
            # print(asin+'-未抓取到品牌')
            # traceback.print_exc()
            pass
        try:  # 解析评分
            score = node.find_element(By.CLASS_NAME, 'a-icon-alt').get_attribute('innerText').split()[0].replace(',', '')
            ws_all.cell(max_row_all + idx, cols['score']).value = score
        except Exception as e:
            print(asin+'-未抓取到评分')
            pass
        try:  # 解析rating数
            temp = node.find_element(By.XPATH, 'div/div/div/div/div/div[2]/div/div/div[2]/div/span[2]')
            rating = int(temp.get_attribute('aria-label').replace(',', ''))
            ws_all.cell(max_row_all + idx, cols['rating']).value = rating
            print(asin+'-抓取到rating：'+ rating)
            break
        except Exception as e:
            try:  # 解析rating数
                temp = node.find_element(By.XPATH, 'div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/span[2]')
                rating = int(temp.get_attribute('aria-label').replace(',', ''))
                ws_all.cell(max_row_all + idx, cols['rating']).value = rating
                print(asin + '-抓取到rating：' + rating)
                break
            except Exception as e:
                pass
        try:  # 解析价格和历史价格
            off_nodes = node.find_elements(By.CLASS_NAME, 'a-offscreen')
            price_gotten = False
            for off_node in off_nodes:
                text = off_node.get_attribute('innerText')
                if text.startswith('$'):
                    value = text.replace('$', '').replace(',', '')
                    if not price_gotten:
                        ws_all.cell(max_row_all + idx, cols['price']).value = value
                        price_gotten = True
                    else:
                        ws_all.cell(max_row_all + idx, cols['price_old']).value = value
        except Exception as e:
            print(asin+'-未抓取到价格')
            pass
        try:  # 解析coupon
            cou = node.find_element(By.CLASS_NAME, 's-coupon-highlight-color')\
                .get_attribute('innerText').replace('Save ', '').replace(' off coupon', '')
            if cou.startswith('$') or cou.endswith('%'):
                ws_all.cell(max_row_all + idx, cols['coupon']).value = cou
        except:
            pass
        all = node.text
        ws_all.cell(max_row_all + idx, 1).value = ky
        ws_all.cell(max_row_all + idx, cols['page']).value = page
        ws_all.cell(max_row_all + idx, cols['index']).value = idx
        ws_all.cell(max_row_all + idx, cols['asin']).value = asin
        c = re.findall(re_coupon, all)
        l = re.findall(re_low_inventory, all)
        low_inventory = -1
        if len(l) > 0:
            low_inventory = l[0]
        ws_all.cell(max_row_all + idx, cols['low_inventory']).value = low_inventory
        if '\nLimited time deal' in all:
            ws_all.cell(max_row_all + idx, cols['deal']).value = 'Limited time deal'
        ws_all.cell(max_row_all + idx, cols['original']).value = all
        if 'Sponsored\n' in all:
            all = all.replace('Sponsored\n', '')
            ws_all.cell(max_row_all + idx, cols['tags']).value = 'Sponsored'
        # if 'Amazon\'s Choice\n' in all:
        #     all = all.replace('Amazon\'s Choice\n', '')
        #     ws_all.cell(max_row_all + idx, cols['tags']).value = 'Amazon\'s Choice'
        for bra in brands_concerned:

            if bra.lower() in all.lower():
                print(asin+' 匹配到品牌 '+bra)
                brand = bra
        ws_all.cell(max_row_all + idx, cols['brand']).value = brand
        all_low = all.lower()
        alls_low = all_low.split('\n')
        prime_price = ''
        for li in alls_low:
            if li.startswith(pri):
                prime_price = li.replace(pri, '').replace(',', '')
        ws_all.cell(max_row_all + idx, cols['prime_price']).value = prime_price

        for ca in capacities_search_order:
            matched = False
            for w in capacities_words[ca]:
                if w in all_low:
                    matched = True
                    break
            if matched:
                ws_all.cell(max_row_all + idx, cols['capacity']).value = ca
                break

        # only_one = True
        for pa in pack_search_order:
            matched = False
            for w in pack_words[pa]:
                if w.lower() in all_low:
                    matched = True
                    break
            if matched:
                ws_all.cell(max_row_all + idx, cols['pack']).value = pa
                # only_one = False
                break
        # if only_one:
        #     ws_all.cell(max_row_all + idx, cols['pack']).value = '1'
        if 'usb 2.0' in all_low or 'usb2.0' in all_low:
            ws_all.cell(max_row_all + idx, cols['interface']).value = '2.0'
        elif 'usb 3.0' in all_low or 'usb3.0' in all_low:
            ws_all.cell(max_row_all + idx, cols['interface']).value = '3.0'
        elif 'usb 3.1' in all_low or 'usb3.1' in all_low:
            ws_all.cell(max_row_all + idx, cols['interface']).value = '3.1'
        elif 'usb 3.2' in all_low or 'usb3.2' in all_low:
            ws_all.cell(max_row_all + idx, cols['interface']).value = '3.2'

        wb.save(file)


for keyword in keywords:
    scrap(keyword)

if len(wb.worksheets) > 1:
    wb.remove(wb['Sheet'])
wb.save(file)
wb.close()
