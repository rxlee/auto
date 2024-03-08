from openpyxl import load_workbook
import os
import time
import json
import random
from os import path
import pymysql

time.sleep(random.randint(5, 10))

''' 数据整理和归档 '''

# today = time.strftime("%Y%m%d", time.localtime())
datapath = 'cap/'

wb_target_all = load_workbook('库存记录表_all.xlsx')
wb_target_yt = load_workbook('库存记录表_Yatong_US_US.xlsx')
wb_target_yhh = load_workbook('库存记录表_Yhh_US_US.xlsx')
wb_target_bc = load_workbook('库存记录表_Baochang_US_US.xlsx')
wb_target_ll = load_workbook('库存记录表_Lanlan_US_US.xlsx')

wb_targets = {"Yatong_US_US": wb_target_yt,
              "Yhh_US_US": wb_target_yhh,
              "Baochang_US_US": wb_target_bc,
              "Lanlan_US_US": wb_target_ll,
              "all": wb_target_all,
              }

target = {'channel': 'channel_name',
    'asin': 'asin',
    'pasin': 'parent_asin',
    'pic': 'thumb_path',
    'title': 'TitleS',
    'time': 'today_time',
    'fulfillable': 'fulfillable_quantity',
    'reserved': 'reserved_quantity',
    'receiving': 'inbound_receiving_quantity',
    'shipped': 'inbound_shipped_quantity',
    'working': 'inbound_working_quantity',
    'total': 'total_quantity',
    'unsellable': 'unsellable_quantity',
    's1': 'yesterday_day_sale',
    's3': '_3_day_sale',
    's7': '_7_day_sale',
    's14': '_14_day_sale',
    's30': '_30_day_sale',
    's60': '_60_day_sale',
    's90': '_90_day_sale',
    'r1': '_1_day_sale_return',
    'r3': '_3_day_sale_return',
    'r7': '_7_day_sale_return',
    'r14': '_14_day_sale_return',
    'r30': '_30_day_sale_return',
    'r60': '_60_day_sale_return',
    'r90': '_90_day_sale_return',
    'a_0_3': '_3_month_age',
    'a_3_6': '_3_6_month_age',
    'a_6_9': '_6_9_month_age',
    'a_9_12': '_9_12_month_age',
    'a_12': '_12_month_age'
}


dir_json = 'cap'

sql1 = '''
INSERT INTO `tb_sku` (`sku`, `channel`, `asin`, `pasin`, `title`, `pic`,`table_name`, `time`, `fulfillable`, `reserved`, `receiving`, `shipped`, `working`
, `total`, `unsellable`, `s1`, `s3`, `s7`, `s14`, `s30`, `s60`, `s90`, `r1`, `r3`, `r7`, `r14`, `r30`, `r60`, `r90`, `a_0_3`, `a_3_6`, `a_6_9`, `a_9_12`, `a_12`) VALUES 
('{sku}', '{channel}', '{asin}', '{pasin}', '{title}', '{pic}', '{table_name}', {time}, {fulfillable}, {reserved}, {receiving}, {shipped}, {working}, 
{total}, {unsellable}, {s1}, {s3}, {s7}, {s14}, {s30}, {s60}, {s90}, {r1}, {r3}, {r7}, {r14}, {r30}, {r60}, {r90}, {a_0_3}, {a_3_6}, {a_6_9}, {a_9_12}, {a_12}) ON DUPLICATE KEY UPDATE 
`asin`='{asin}',`pasin`='{pasin}',`title`='{title}',`pic`='{pic}',`table_name`='{table_name}',`time`={time},`fulfillable`={fulfillable},`reserved`={reserved},`receiving`={receiving},`shipped`={shipped},`working`={working}
,`total`={total},`unsellable`={unsellable},`s1`={s1},`s3`={s3},`s7`={s7},`s14`={s14},`s30`={s30},`s60`={s60},`s90`={s90}
,`r1`={r1},`r3`={r3},`r7`={r7},`r14`={r14},`r30`={r30},`r60`={r60},`r90`={r90},`a_0_3`={a_0_3},`a_3_6`={a_3_6},`a_6_9`={a_6_9},`a_9_12`={a_9_12},`a_12`={a_12};
'''


sql2 = '''
INSERT INTO `{table_name}` (`time`, `fulfillable`, `reserved`, `receiving`, `shipped`, `working`
, `total`, `unsellable`, `s1`, `s3`, `s7`, `s14`, `s30`, `s60`, `s90`, `r1`, `r3`, `r7`, `r14`, `r30`, `r60`, `r90`, `a_0_3`, `a_3_6`, `a_6_9`, `a_9_12`, `a_12`) VALUES 
({time}, {fulfillable}, {reserved}, {receiving}, {shipped}, {working}, 
{total}, {unsellable}, {s1}, {s3}, {s7}, {s14}, {s30}, {s60}, {s90}, {r1}, {r3}, {r7}, {r14}, {r30}, {r60}, {r90}, {a_0_3}, {a_3_6}, {a_6_9}, {a_9_12}, {a_12}) ON DUPLICATE KEY UPDATE 
`fulfillable`={fulfillable},`reserved`={reserved},`receiving`={receiving},`shipped`={shipped},`working`={working}
,`total`={total},`unsellable`={unsellable},`s1`={s1},`s3`={s3},`s7`={s7},`s14`={s14},`s30`={s30},`s60`={s60},`s90`={s90}
,`r1`={r1},`r3`={r3},`r7`={r7},`r14`={r14},`r30`={r30},`r60`={r60},`r90`={r90},`a_0_3`={a_0_3},`a_3_6`={a_3_6},`a_6_9`={a_6_9},`a_9_12`={a_9_12},`a_12`={a_12};
'''

sql3 = '''
CREATE TABLE `{table_name}` like tb_template;
'''

try:
    db = pymysql.connect(host='127.0.0.1', user='root', passwd='123456', port=3306, db='dj')
    print('连接成功！')
except:
    print('链接失败，检查MySQL服务!')

cursor = db.cursor()

d_tables = []
cursor.execute("SHOW TABLES LIKE 'd%';")
    # 获取所有记录列表
results = cursor.fetchall()
for line in results:
    d_tables.append(line[0])
print(d_tables)

files = os.listdir(dir_json)
files = sorted(files, reverse=False)

for file in files:
    data = json.load(open(path.join(datapath, file)))
    if len(data) > 0:
        for sku in data.keys():
            try:
                if sku != 'time' and sku != 'keywords':
                    table_name = 'd_' + sku + '_' + data[sku]['channel_name']
                    table_name = table_name.lower()
                    tg = {'sku': sku, 'table_name': table_name}
                    if table_name not in d_tables:
                        try:
                            cursor.execute(sql3.format(**tg))  # 预处理新SKU的历史数据表
                            db.commit()                        # 必须提交好
                        except:
                            print(table_name+'创建历史数据表失败！')
                            continue
                    else:
                        print(table_name+'历史数据表已存在，无需创建~')
                    # 必须提交好
                    for key in target.keys():
                        tg[key] = data[sku][target[key]]
                    print(sql1.format(**tg))
                    print(sql3.format(**tg))
                    cursor.execute(sql1.format(**tg)) ## 数据信息表更新
                    cursor.execute(sql2.format(**tg)) ## 历史数据表更新
                    db.commit()
            except:
                # 如果发生错误则回滚
                db.rollback()
                print(sku+'---数据插入错误！')
                continue
            # break
    # break



# 关闭数据库连接
db.close()



# for file in files:
#     data = json.load(open(path.join(datapath, file)))  ## 目录已经换了
#     if len(data) > 0:
#         for key in data.keys():
#             print(key)
        # for wb_target in wb_targets.values():
        #     for s in target:  # 提取所有格式化数据
        #         update_sheet(wb_target, file, s.get('sheet'), data, s.get('key'))
        #     update_oos(wb_target, file, data)  # 计算断货与否

