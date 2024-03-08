from openpyxl import load_workbook
import os
import time
import json
import random


time.sleep(random.randint(5, 10))

'''
'''

today = time.strftime("%Y-%m-%d", time.localtime())
jsonpath = 'cap/' + str(today).replace('-', '')

ex_ds = load_workbook('库存记录表.xlsx')
ex_tg = load_workbook('库存健康.xlsx')

# 计算最近30天，断货天数，乘以系数
ratios_oos = {}
st_ds_ratio = ex_ds['容差系数']
for i in range(2, st_ds_ratio.max_row + 1):
    ratio = float(st_ds_ratio.cell(i, 2).value)
    ratios_oos[st_ds_ratio.cell(i, 1).value] = ratio

# 计算30天内的在售天数----减去的断货天数已经乘了容差系数
days_sell = {}
st_ds_oos = ex_ds['断货记录']
oos_max_col = st_ds_oos.max_column
print('断货天数统计周期：从【'+st_ds_oos.cell(1, oos_max_col-29).value+'】至【'+st_ds_oos.cell(1, oos_max_col).value+'】')
for i in range(2, st_ds_oos.max_row + 1):
    day_oos = 0
    for j in range(oos_max_col-29, oos_max_col + 1):
        day_oos += st_ds_oos.cell(i, j).value
    sku = st_ds_oos.cell(i, 1).value
    ratio = 0.8 if sku not in ratios_oos else ratios_oos.get(sku)
    days_sell[sku] = 30 - (30 - day_oos) * ratio
    # print(sku+'\t'+str(day_oos)+'\t'+str(ratio)+'\t'+ str(days_sell[sku]))

# 计算申请中的货件数量
q_requesting = {}
st_ds_req = ex_ds['申请中']
for i in range(2, st_ds_req.max_row + 1):
    sku = st_ds_req.cell(i, 1).value
    if sku is None or sku.strip() == '':
        continue
    n = st_ds_req.cell(i, 2).value
    if n is None or str(n).strip() == '':
        continue
    num = n if sku not in q_requesting else q_requesting[sku] + n
    q_requesting[sku] = num

# for i in requestings:
#     print(i, requestings[i])

# 获取最新一列的数据
def latest(day):
    latest = {}
    st = ex_ds[str(day)]
    for i in range(2, st.max_row + 1):
        latest[st.cell(i, 1).value] = st.cell(i, st.max_column).value
    return latest


sell_7 = latest(7)
sell_14 = latest(14)
sell_30 = latest(30)
FBA_total = latest('total')
FBA_unsellable = latest('不可售')

# 同步备货方案一 S30+(S7+S14)/4
sell_day = {}
for sku in sell_30:
    q_requesting[sku] = 0 if sku not in q_requesting else q_requesting[sku]  # 顺便把请求中的0数据初始化一下
    sell_day[sku] = 0 if days_sell[sku] == 0 else (sell_30[sku] + (sell_7[sku] + sell_14[sku]) / 4) / days_sell[sku]



def update_sheet(sheet, d, sep, d_only = False):  # 原样插入d，不用除以日均（用于插入原日均值）
    ws = ex_tg[sheet]
    last = ws.cell(row=1, column=ws.max_column).value
    write_col = ws.max_column
    if today != last:  # 新的一天，往后新写一列
        write_col += 1
    ws.cell(row=1, column=write_col).value = today

    for i in range(2, ws.max_row + 1):
        sku = ws.cell(i, 1).value
        if sku in d:
            q = d[sku]
            s = sep[sku]
            res = 0
            if d_only:
                res = q
            else:
                if q == 0:
                    res = 0
                else:
                    if s == 0:
                        res = 999
                    else:
                        res = q / s
            ws.cell(row=i, column=write_col).value = round(res, 3 if d_only else 1)


def append(all, other):
    for sku in all:
        all[sku] = all[sku] + other[sku]


q_sellable = latest('可售')
q_reserve = latest('预留')
q_receiving = latest('接收')
q_shipped = latest('shipped')
q_working = latest('working')

update_sheet('日均销量', sell_day, sell_day, d_only=True)
update_sheet('可售', q_sellable, sell_day)
update_sheet('预留', q_reserve, sell_day)
update_sheet('接收', q_receiving, sell_day)
update_sheet('shipped', q_shipped, sell_day)
update_sheet('working', q_working, sell_day)
update_sheet('申请中', q_requesting, sell_day)

q_appending = q_sellable

append(q_appending, q_reserve)
update_sheet('可+预', q_appending, sell_day)

append(q_appending, q_receiving)
update_sheet('可+预+接', q_appending, sell_day)

append(q_appending, q_shipped)
update_sheet('可+预+接+s', q_appending, sell_day)

append(q_appending, q_working)
update_sheet('可_TO_w', q_appending, sell_day)

append(q_appending, q_requesting)
update_sheet('可_TO_申', q_appending, sell_day)


ex_tg.save('库存健康.xlsx')
