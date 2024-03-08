from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import time
import os
from os import path
import random

dir_json = 'keywords'

wb = load_workbook('关键词.xlsx')
keywords = []
ws = wb['备选关键词']
for i in range(2, ws.max_row+1):
    keyword = ws.cell(i, 1).value
    if keyword is None or keyword.strip() == '':
        continue
    keywords.append(keyword)

files = os.listdir(dir_json)
files = sorted(files)[::-1]
file_latest = files[0]
data = json.load(open(path.join(dir_json, file_latest)))
asins = []
titles = {}
for i in data['data']:
    asins.append(i['id'])
    titles[i['id']] = i['title'] if 'title' in i else ''

riqis = []
ds = {}
for f in files:
    data = json.load(open(path.join(dir_json, f)))
    dt = f.split(' ')[0]
    riqis.append(dt)
    # print(dt)
    dall = {}
    for i in data['data']:
        d = {}
        bsr = int(i['metadataMap']['render.zg.rank'])
        row = bsr + 1
        asin = i['id']
        d['id'] = asin
        if asin not in asins:
            asins.append(asin)
            if 'title' in i:
                titles[asin] = i['title']
            else:
                titles[asin] = ''
        d['bsr'] = bsr
        if 'bsr' not in i:
            d['score'] = -1
            d['rating'] = -1
            d['price'] = -1
            dall[asin] = d
            continue
        d['score'] = i['score']
        d['rating'] = i['rating']
        d['price'] = i['price']
        d['title'] = i['title']
        dall[asin] = d
    ds[dt] = dall
    # print(real_url)

riqis = riqis[::-1]

def upd(key):
    sheet = wb[key]
    sheet.cell(1, 1).value = 'asin'
    sheet.cell(1, 2).value = 'title'
    for idx, a in enumerate(asins):
        sheet.cell(idx + 2, 1).value = a
        sheet.cell(idx + 2, 2).value = titles[a]
    col = 3
    for riqi in riqis:
        d = ds[riqi]
        sheet.cell(1, col).value = riqi
        print(d)
        for row in range(2, len(asins) + 2):
            a = sheet.cell(row, 1).value
            if a in d.keys():
                sheet.cell(row, col).value = d[a][key]
                # print(d[a]['bsr'])
        col += 1

upd('bsr')
upd('score')
upd('rating')
upd('price')






wb.save('C:/Users/Administrator/Desktop/bsr.xlsx')
